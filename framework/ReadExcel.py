from openpyxl import load_workbook
from framework.logger import Logger
import re
import configparser,os
proDir = os.getcwd()
configPath = os.path.join(proDir, "config\config.ini")
cf = configparser.ConfigParser()

cf.read(configPath)#encoding="utf-8-sig"
#print(cf.get("Data", "1")
logger = Logger(logger="RWExcel").getlog()
class RExcetodicl():

    def RExcetodicl(self,path_ex):
        wb = load_workbook(path_ex)
        # 获取所有表名
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
        # 根据表名打开sheet表
        sheet1 = wb[sheet_names[0]]  # 打开第一个 sheet 工作表
        # 获取C列的所有数据
        datas=[]
        pattern = re.compile(u'^((https|http|ftp|rtsp|mms)?:\/\/)[^\s]+')
        for b, e, c, a, d,f,g,h,i in zip(sheet1["B"][1:], sheet1["E"][1:], sheet1["C"][1:], sheet1["A"][1:],sheet1["D"][1:], sheet1["F"][1:],sheet1["G"][1:],sheet1["H"][1:],sheet1["I"][1:]):

            if (pattern.search(d.value))==None:#判断读取的地址是否有前缀地址
                url='%s%s'%(cf.get("Data", "address"),d.value)#无前缀读取配置文件添加前缀
            else:
                url=d.value#有前缀使用读取的完整地址
            data={
                "caseID":a.value,
                "url":url,
                "param": g.value,
                "method": e.value,
                "headers": f.value,
                "teststep":c.value,
                "testname": b.value,
                "expect": h.value,
                "case_class": i.value,
                "path_ex":path_ex,
                "row":a.value,
            }
            #logger.info(data)
            if data["case_class"]=="K":
                datas.append(data)

        try:
            wb.save(path_ex)
            #logger.info('测试数据保存成功！！！')
        except Exception as e:
            logger.error(e)
            logger.error('保存失败，可能Excel文件未关闭，请关闭Excel文件后重新测试')
        #logger.info(datas)
        return  {
                "message": "ok",
                "status": "1000",
                "data": datas
            }






