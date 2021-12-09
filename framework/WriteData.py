from framework.FindValue import FindValue
from framework.logger import Logger
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
logger = Logger(logger="WriteData").getlog()
class WriteData():

    def writedata(self,PM,expect,sheet1,row,testname):

        #PM['Response_Data'] = PM['Response_Data'] .replace("'", '"')  # 替换"'", '"'
        #dict_data = json.loads(PM['Response_Data'] )  # 转python字典

        if FindValue().find_value(PM['Response_Data'] , expect) == True:  # ReadAPI.get(url, param,testname)[3])==200:

            sheet1.cell(row=row, column=7, value=str(PM['param']))
            sheet1.cell(row=row, column=10, value=str(PM['Response_Data']))  # 响应结果
            sheet1.cell(row=row, column=11, value=PM['time'])  # 请求时间
            sheet1.cell(row=row, column=12, value=int(PM['Status_Code']))  # 状态码

            #sheet1.cell(row=row, column=12, value="pass")  # 判断通过

            # 判断通过
            Color = ['c6efce','006100']#绿
            fille = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色
            font = Font(u'宋体', size=12, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=row, column=13, value="").fill = fille  # 序列
            sheet1.cell(row=row, column=13, value="PASS").font = font  # 序列
            # 判断通过
            #logger.info('《' + str(testname) + '》项，响应成功、响应时间：' + str(PM['time']) + '、状态码：' + str(PM['Status_Code']))
            return "pass"

        else:
            sheet1.cell(row=row, column=10, value=str(PM['Response_Data']))  # 响应结果
            sheet1.cell(row=row, column=11, value=PM['time'])  # 请求时间
            sheet1.cell(row=row, column=12, value=int(PM['Status_Code']))  # 状态码
            #sheet1.cell(row=row, column=12, value="fail")  # 判断失败

            # 判断失败
            Color = ['ffc7ce', '9c0006']#红
            fille = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色
            font = Font(u'宋体', size=12, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=row, column=13, value="").fill = fille  # 序列
            sheet1.cell(row=row, column=13, value="FAIL").font = font  # 序列
            # 判断失败

            #logger.info('《' + str(testname) + '》项，响应成功、响应时间：' + str(PM['time']) + '、状态码：' + str(PM['Status_Code']))
            return "fail"



